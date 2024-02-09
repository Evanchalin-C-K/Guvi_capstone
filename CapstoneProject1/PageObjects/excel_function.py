from openpyxl import load_workbook


class Excel_Function:
    def __init__(self, file_name, sheet_name):
        self.file = file_name
        self.sheet = sheet_name

    # column count
    def column_count(self):
        workbook = load_workbook(self.file)
        sheet = workbook[self.sheet]
        return sheet.max_column

    def row_count(self):
        workbook = load_workbook(self.file)
        sheet = workbook[self.sheet]
        return sheet.max_row

    def read_file(self, row_number, column_number):
        workbook = load_workbook(self.file)
        sheet = workbook[self.sheet]
        return (sheet.cell(row_number, column_number)).value

    def write_file(self, row_number, column_number, data):
        workbook = load_workbook(self.file)
        sheet = workbook[self.sheet]
        (sheet.cell(row_number, column_number)).value = data
        workbook.save(self.file)
